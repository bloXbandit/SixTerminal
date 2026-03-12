import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd
import logging

logger = logging.getLogger(__name__)

class DashboardGenerator:
    """
    The Artist 🎨
    Takes analyzed data and paints it into a professional Excel dashboard.
    Now includes REAL Stairway Charts!
    """

    def __init__(self, analyzer, output_path: str):
        self.analyzer = analyzer
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        
        # Define Styles
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="36454F", end_color="36454F", fill_type="solid") # Charcoal
        self.border_thin = Side(border_style="thin", color="000000")
        self.border_all = Border(left=self.border_thin, right=self.border_thin, top=self.border_thin, bottom=self.border_thin)

    def generate(self):
        """Main orchestration method to build all sheets."""
        logger.info("Generating Dashboard...")
        
        # 1. Executive Summary
        self._build_executive_summary()
        
        # 2. Stairway / Milestone Tracker (Now with Charts!)
        self._build_stairway_tracker()
        
        # 3. Critical Path View
        self._build_critical_path_sheet()
        
        # 4. Procurement Log
        self._build_procurement_sheet()

        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            del self.wb["Sheet"]

        try:
            self.wb.save(self.output_path)
            logger.info(f"Dashboard saved to: {self.output_path}")
            return self.output_path
        except Exception as e:
            logger.error(f"Failed to save Excel: {e}")
            raise

    def _build_executive_summary(self):
        ws = self.wb.create_sheet("Executive Summary", 0)
        ws.sheet_view.showGridLines = False
        
        ws["B2"] = "PROJECT EXECUTIVE DASHBOARD"
        ws["B2"].font = Font(size=24, bold=True, color="333333")
        
        stats = self.analyzer.get_dashboard_summary()
        
        metrics = [
            ("Data Date", stats.get("data_date", "N/A")),
            ("Total Activities", stats.get("total_activities", 0)),
            ("Critical Activities", stats.get("critical_activities", 0)),
            ("Slipping (>5 days)", stats.get("slipping_activities", 0)),
            ("% Critical", f"{stats.get('percent_critical', 0)}%")
        ]
        
        row = 4
        for label, value in metrics:
            ws.cell(row=row, column=2, value=label).font = Font(bold=True)
            ws.cell(row=row, column=3, value=value)
            row += 1
            
        if stats.get("percent_critical", 0) > 20:
            ws["E4"] = "⚠️ HIGH RISK: Critical Path Saturation"
            ws["E4"].font = Font(color="FF0000", bold=True, size=14)

        self._autofit_columns(ws)

    def _build_stairway_tracker(self):
        ws = self.wb.create_sheet("Milestone Stairway")
        
        df = self.analyzer.get_milestones()
        if df.empty:
            ws["A1"] = "No Milestones Found"
            return

        # Prepare Data Table
        cols = ['task_code', 'task_name', 'target_end_date', 'current_finish', 'variance_days', 'status_readable']
        headers = ['ID', 'Milestone Name', 'Baseline Date', 'Forecast Date', 'Variance', 'Status']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row_data in enumerate(df[cols].itertuples(index=False), 2):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border_all
                if c_idx in [3, 4] and pd.notnull(value):
                    cell.number_format = 'mm/dd/yyyy'
                
                # Conditional Formatting logic
                if c_idx == 5:
                    if value > 10: cell.fill = PatternFill(start_color="FF9999", fill_type="solid")
                    elif value > 0: cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")
                    else: cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")

        self._autofit_columns(ws)
        
        # --- BUILD THE STAIRWAY CHART ---
        # A scatter chart comparing Baseline vs Forecast
        chart = ScatterChart()
        chart.title = "Milestone Stairway (Baseline vs Forecast)"
        chart.style = 13
        chart.x_axis.title = 'Date'
        chart.y_axis.title = 'Milestone Sequence'
        chart.height = 15
        chart.width = 25

        last_row = len(df) + 1
        
        # Series 1: Forecast Dates (The Actual Stairway)
        # Y values (just 1, 2, 3... to stack them vertically)
        # X values (Dates)
        
        # Note: Excel Charts are tricky with Dates programmatically sometimes.
        # We define ranges.
        
        x_values = Reference(ws, min_col=4, min_row=2, max_row=last_row) # Forecast Date
        # We need a dummy sequence column for Y-axis to make them "Step up"
        # Let's add a hidden helper column G with 1, 2, 3...
        for i in range(2, last_row + 1):
            ws.cell(row=i, column=7, value=i-1)
            
        y_values = Reference(ws, min_col=7, min_row=2, max_row=last_row) # Sequence
        
        series_forecast = Series(y_values, x_values, title="Forecast Date")
        series_forecast.marker.symbol = "circle"
        series_forecast.marker.graphicalProperties.solidFill = "0000FF" # Blue dots
        chart.series.append(series_forecast)
        
        # Series 2: Baseline Dates (The Plan)
        x_values_base = Reference(ws, min_col=3, min_row=2, max_row=last_row)
        series_base = Series(y_values, x_values_base, title="Baseline Date")
        series_base.marker.symbol = "diamond"
        series_base.marker.graphicalProperties.solidFill = "808080" # Grey dots
        chart.series.append(series_base)

        ws.add_chart(chart, "I2") # Place chart next to table

    def _build_critical_path_sheet(self):
        ws = self.wb.create_sheet("Critical Path")
        df = self.analyzer.get_critical_path()
        if df.empty:
            ws["A1"] = "No Critical Path Activities Found"
            return

        cols = ['task_code', 'task_name', 'current_start', 'current_finish', 'total_float_hr_cnt']
        headers = ['Activity ID', 'Name', 'Start', 'Finish', 'Total Float (Hrs)']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for r_idx, row_data in enumerate(df[cols].itertuples(index=False), 2):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if c_idx in [3, 4] and pd.notnull(value):
                    cell.number_format = 'mm/dd/yyyy'

        self._autofit_columns(ws)

    def _build_procurement_sheet(self):
        ws = self.wb.create_sheet("Procurement Log")
        df = self.analyzer.get_procurement_log()
        if df.empty:
            ws["A1"] = "No Procurement Items Found"
            return
            
        cols = ['task_code', 'task_name', 'current_finish', 'status_readable']
        headers = ['ID', 'Item Description', 'Date Required', 'Status']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        for r_idx, row_data in enumerate(df[cols].itertuples(index=False), 2):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if c_idx == 3 and pd.notnull(value):
                    cell.number_format = 'mm/dd/yyyy'

        self._autofit_columns(ws)

    def _autofit_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)
